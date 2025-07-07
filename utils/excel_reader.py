"""
utils/excel_reader.py  v2.0 -- 2025 update
Added get_row_count(), get_sheet_names(), improved error handling.
"""
import openpyxl, os
from framework.logger import get_logger

logger = get_logger(__name__)


class ExcelReader:
    """Reads test data from Excel (.xlsx) files using openpyxl."""

    def __init__(self, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        self.file_path = file_path
        self.workbook  = openpyxl.load_workbook(file_path)
        logger.info(f"Loaded Excel: {file_path}")

    def get_sheet_names(self) -> list:
        return self.workbook.sheetnames

    def get_row_count(self, sheet_name: str) -> int:
        return self.workbook[sheet_name].max_row - 1  # exclude header

    def get_all_rows(self, sheet_name: str) -> list:
        sheet   = self.workbook[sheet_name]
        rows    = list(sheet.iter_rows(values_only=True))
        if not rows: return []
        headers = [str(h).strip() for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]
                if any(c is not None for c in row)]

    def get_column_data(self, sheet_name: str, col: int) -> list:
        sheet = self.workbook[sheet_name]
        return [sheet.cell(row=r, column=col).value
                for r in range(2, sheet.max_row + 1)]

    def close(self):     self.workbook.close()
    def __enter__(self): return self
    def __exit__(self, *args): self.close()


def load_sheet(file_path: str, sheet_name: str) -> list:
    """One-liner convenience function for test files."""
    with ExcelReader(file_path) as r:
        return r.get_all_rows(sheet_name)
